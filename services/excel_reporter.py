# -*- coding: utf-8 -*-
"""
@file: excel_reporter.py
@description: Генерация Excel-отчёта посещаемости из плоских данных БД (без шаблона).
  Формат: pivot-таблица — строки=участники, столбцы=даты тренировок, ячейки=1/пусто.
@dependencies: openpyxl, config
@created: 2025-02-25
"""

from __future__ import annotations

import logging
from datetime import date as date_type, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger("rzdbadminton")

MONTHS_RU = {
    1: "Январь", 2: "Февраль", 3: "Март", 4: "Апрель", 5: "Май", 6: "Июнь",
    7: "Июль", 8: "Август", 9: "Сентябрь", 10: "Октябрь", 11: "Ноябрь", 12: "Декабрь",
}

_COLOR_HEADER = "2E5FA3"
_COLOR_SUBHEADER = "C5D9F1"
_COLOR_ALT_ROW = "EEF4FB"
_COLOR_TOTAL_ROW = "BDD7EE"
_COLOR_WHITE = "FFFFFF"
_COLOR_DARK = "1F3864"
_COLOR_ACCENT = "2E5FA3"


def _side() -> Side:
    return Side(style="thin", color="9DC3E6")


def _border() -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _build_pivot(
    records: list[tuple[int, str, str, date_type]],
) -> tuple[
    list[tuple[int, str, str]],
    list[date_type],
    dict[tuple[int, date_type], bool],
]:
    """
    Построить pivot из плоских записей посещаемости.

    Args:
        records: список (user_id, display_name, full_name, poll_date).

    Returns:
        users: отсортированный список (user_id, display_name, full_name).
        training_dates: отсортированный список дат тренировок.
        attendance: словарь {(user_id, poll_date): True} для присутствовавших.
    """
    user_map: dict[int, tuple[str, str]] = {}
    dates: set[date_type] = set()
    attendance: dict[tuple[int, date_type], bool] = {}

    for user_id, display_name, full_name, poll_date in records:
        if user_id not in user_map:
            user_map[user_id] = (display_name, full_name)
        dates.add(poll_date)
        attendance[(user_id, poll_date)] = True

    users = sorted(
        [(uid, dm[0], dm[1]) for uid, dm in user_map.items()],
        key=lambda x: (x[2] or x[1]).lower(),
    )
    return users, sorted(dates), attendance


def _name_label(display_name: str, full_name: str) -> str:
    """Сформировать отображаемое имя: «Имя Фамилия (@username)» или просто одно из двух."""
    dn = display_name.lstrip("@").replace("_", " ").lower()
    fn = full_name.lower()
    if full_name and display_name and dn != fn:
        return f"{full_name} ({display_name})"
    return full_name or display_name


def _write_excel(
    records: list[tuple[int, str, str, date_type]],
    report_date: date_type,
    output_path: Path,
) -> None:
    """
    Сгенерировать Excel-файл из плоских данных посещаемости.

    Args:
        records: список (user_id, display_name, full_name, poll_date).
        report_date: дата отчёта (используется для заголовка и имени листа).
        output_path: путь сохранения файла.
    """
    month_name = MONTHS_RU.get(report_date.month, str(report_date.month))
    title = f"Посещаемость · Бадминтон РЖД · {month_name} {report_date.year}"

    users, training_dates, attendance = _build_pivot(records)
    n_dates = len(training_dates)
    n_cols = 2 + n_dates + 1  # №, ФИО/Ник, [dates...], Итого

    wb = Workbook()
    ws = wb.active
    ws.title = f"{month_name} {report_date.year}"

    border = _border()

    # ── Row 1: заголовок (объединённые ячейки) ─────────────────────────────
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(bold=True, size=13, color=_COLOR_WHITE)
    title_cell.fill = _fill(_COLOR_HEADER)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # ── Row 2: заголовки столбцов ──────────────────────────────────────────
    DAYS_RU = {0: "Пн", 1: "Вт", 2: "Ср", 3: "Чт", 4: "Пт", 5: "Сб", 6: "Вс"}
    date_headers = [f"{DAYS_RU.get(d.weekday(), '')}\n{d.strftime('%d.%m')}" for d in training_dates]
    headers = ["№", "ФИО / Ник"] + date_headers + ["Итого"]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font = Font(bold=True, size=10, color=_COLOR_DARK)
        cell.fill = _fill(_COLOR_SUBHEADER)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[2].height = 34

    # ── Rows 3+: данные участников ─────────────────────────────────────────
    for row_idx, (user_id, display_name, full_name) in enumerate(users, start=1):
        r = row_idx + 2
        row_fill = _fill(_COLOR_ALT_ROW) if row_idx % 2 == 0 else None

        def _cell(col: int, value=None):  # type: ignore[return]
            c = ws.cell(row=r, column=col, value=value)
            c.border = border
            if row_fill:
                c.fill = row_fill
            return c

        num_c = _cell(1, row_idx)
        num_c.alignment = Alignment(horizontal="center", vertical="center")

        name_c = _cell(2, _name_label(display_name, full_name))
        name_c.alignment = Alignment(horizontal="left", vertical="center")

        total = 0
        for c_idx, d in enumerate(training_dates, start=3):
            att_cell = _cell(c_idx)
            if attendance.get((user_id, d)):
                att_cell.value = 1
                att_cell.font = Font(bold=True, color=_COLOR_ACCENT)
                total += 1
            att_cell.alignment = Alignment(horizontal="center", vertical="center")

        итого_c = _cell(2 + n_dates + 1, total)
        итого_c.font = Font(bold=True)
        итого_c.alignment = Alignment(horizontal="center", vertical="center")

    # ── Итоговая строка: кол-во присутствовавших на каждой тренировке ──────
    if users:
        sum_row = len(users) + 3
        label_c = ws.cell(row=sum_row, column=2, value="Всего присутствовало:")
        label_c.font = Font(bold=True, size=10)
        label_c.alignment = Alignment(horizontal="right", vertical="center")
        label_c.border = border
        label_c.fill = _fill(_COLOR_TOTAL_ROW)

        num_c = ws.cell(row=sum_row, column=1)
        num_c.border = border
        num_c.fill = _fill(_COLOR_TOTAL_ROW)

        total_all = 0
        for c_idx, d in enumerate(training_dates, start=3):
            count = sum(1 for uid, _, _ in users if attendance.get((uid, d)))
            total_all += count
            c = ws.cell(row=sum_row, column=c_idx, value=count)
            c.font = Font(bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = border
            c.fill = _fill(_COLOR_TOTAL_ROW)

        итого_c = ws.cell(row=sum_row, column=2 + n_dates + 1, value=total_all)
        итого_c.font = Font(bold=True)
        итого_c.alignment = Alignment(horizontal="center", vertical="center")
        итого_c.border = border
        итого_c.fill = _fill(_COLOR_TOTAL_ROW)
        ws.row_dimensions[sum_row].height = 20

    # ── Ширина столбцов ────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 32
    for c_idx in range(3, 3 + n_dates):
        ws.column_dimensions[get_column_letter(c_idx)].width = 9
    if n_dates:
        ws.column_dimensions[get_column_letter(2 + n_dates + 1)].width = 9

    # ── Закрепить строки заголовка ─────────────────────────────────────────
    ws.freeze_panes = "C3"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info(
        "Excel-отчёт сохранён: %s · %s участников · %s тренировок",
        output_path, len(users), n_dates,
    )


async def get_report_file(
    report_date: date_type | datetime,
    records: list[tuple[int, str, str, date_type]],
) -> str | None:
    """
    Сформировать Excel-отчёт посещаемости за месяц из плоских данных.

    Args:
        report_date: дата отчёта (определяет месяц заголовка и имя файла).
        records: список (user_id, display_name, full_name, poll_date) из get_monthly_attendance_records().

    Returns:
        Путь к созданному .xlsx файлу или None при ошибке.
    """
    if isinstance(report_date, datetime):
        report_date = report_date.date()

    try:
        output_path = Path("data/reports") / f"attendance_{report_date.year}_{report_date.month:02d}.xlsx"
        _write_excel(records, report_date, output_path)
        return str(output_path)
    except Exception as e:
        logger.exception("Ошибка формирования Excel-отчёта: %s", e)
        return None
