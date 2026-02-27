"""
@file: report.py
@description: Локальное формирование отчёта — Doc/Бадминтон.xlsx, листы по месяцам
@dependencies: openpyxl, config
@created: 2026-02-25
"""

import logging
import re
import tempfile
import zipfile
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from config import get_settings

logger = logging.getLogger("rzdbadminton")

# Стили границ, поддерживаемые openpyxl (Excel OOXML)
VALID_BORDER_STYLES = frozenset({
    "slantDashDot", "mediumDashed", "double", "dashDotDot", "medium", "thin",
    "hair", "mediumDashDotDot", "thick", "dashed", "dashDot", "mediumDashDot", "dotted",
})

MONTHS_RU = {
    1: "январь", 2: "февраль", 3: "март", 4: "апрель", 5: "май", 6: "июнь",
    7: "июль", 8: "август", 9: "сентябрь", 10: "октябрь", 11: "ноябрь", 12: "декабрь",
}


def _get_report_path() -> str:
    """Путь к файлу отчёта."""
    return str(get_settings().report_file)


def _fix_stylesheet_in_xlsx(xlsx_path: str) -> None:
    """
    Исправить несовместимые стили границ в xlsx (Excel/Google Sheets добавляют стили,
    которые openpyxl не поддерживает). Модифицирует файл на месте.
    """
    path = Path(xlsx_path)
    with tempfile.TemporaryDirectory() as tmp:
        extract_dir = Path(tmp) / "xlsx"
        with zipfile.ZipFile(path, "r") as z:
            z.extractall(extract_dir)
        styles_path = extract_dir / "xl" / "styles.xml"
        if not styles_path.exists():
            return
        content = styles_path.read_text(encoding="utf-8")

        def replace_invalid_border_style(match: re.Match) -> str:
            val = match.group(1).strip()
            if val not in VALID_BORDER_STYLES:  # пустой, неизвестный или невалидный
                return 'style="thin"'
            return match.group(0)

        # Заменяем невалидные style= во всём файле (в styles.xml это в основном границы)
        new_content = re.sub(r'\bstyle="([^"]*)"', replace_invalid_border_style, content)
        if new_content == content:
            return
        styles_path.write_text(new_content, encoding="utf-8")
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
            for f in sorted(extract_dir.rglob("*")):
                if f.is_file():
                    arcname = f.relative_to(extract_dir)
                    z.write(f, arcname)
        logger.info("Исправлены несовместимые стили в %s", xlsx_path)


def _load_workbook_safe(path: str):
    """Загрузить workbook, при ошибке стилей — исправить и повторить."""
    try:
        return load_workbook(path)
    except ValueError as e:
        err = str(e).lower()
        if "stylesheet" in err or "invalid" in err or "border" in err or "style" in err:
            logger.warning("Файл %s не читается openpyxl (стили): %s. Исправляю...", path, e)
            _fix_stylesheet_in_xlsx(path)
            try:
                return load_workbook(path)
            except ValueError as e2:
                logger.warning("Исправление не помогло. Создаю новый файл, старый → %s.backup", path)
                path_obj = Path(path)
                backup = path_obj.with_suffix(path_obj.suffix + ".backup")
                path_obj.rename(backup)
                _ensure_report_file(path)
                return load_workbook(path)
        raise


def _sheet_name_for_month(year: int, month: int) -> str:
    """Имя листа для месяца: «Январь 2025»."""
    name = MONTHS_RU.get(month, str(month))
    return f"{name} {year}"


def _ensure_report_file(path: str) -> bool:
    """Создать файл: копия шаблона или минимальная структура."""
    import shutil
    p = Path(path)
    if p.exists():
        return True
    p.parent.mkdir(parents=True, exist_ok=True)
    template = Path(get_settings().report_template_file)
    if template.exists():
        shutil.copy2(template, path)
        _fix_stylesheet_in_xlsx(path)  # совместимость с openpyxl
        logger.info("Создан отчёт из шаблона: %s", path)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Шаблон"
        ws["A1"] = "Дата"
        ws["B1"] = "ФИ"
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 30
        wb.save(path)
        logger.info("Создан шаблон отчёта (шаблон не найден): %s", path)
    return True


def _is_formula(val) -> bool:
    """Проверить, что значение — формула Excel."""
    return isinstance(val, str) and val.strip().startswith("=")


def _copy_formula_row(formula: str, from_row: int, to_row: int) -> str:
    """Скопировать формулу, заменив ссылки на строку from_row на to_row."""
    if from_row == to_row:
        return formula

    def repl(match):
        col_part, row_part = match.group(1), match.group(2)
        if row_part == str(from_row):
            return col_part + str(to_row)
        return match.group(0)

    return re.sub(r"(\$?[A-Z]+\$?)(\d+)", repl, formula)


def _get_formula_template_row(ws) -> int | None:
    """Найти первую строку с формулами (обычно строка 2) для копирования."""
    for row in range(2, min(ws.max_row + 1, 10)):
        for col in range(3, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if _is_formula(cell.value):
                return row
    return None


def _get_or_create_month_sheet(wb: Workbook, report_date: date):
    """Получить или создать лист для месяца (полная копия структуры первого листа)."""
    sheet_name = _sheet_name_for_month(report_date.year, report_date.month)
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    template = wb.active
    new_ws = wb.create_sheet(title=sheet_name)
    max_col = max(template.max_column, 2)
    for col in range(1, max_col + 1):
        src_cell = template.cell(row=1, column=col)
        default = ("Дата", "ФИ")[col - 1] if col <= 2 else ""
        new_ws.cell(row=1, column=col, value=src_cell.value or default)
        col_letter = get_column_letter(col)
        cd = template.column_dimensions.get(col_letter)
        new_ws.column_dimensions[col_letter].width = (cd.width if cd else None) or (12 if col == 1 else 30)
    return new_ws


def _append_report_data(ws, template_ws, report_date: date, names: list[str]) -> None:
    """Добавить данные за дату в лист, сохраняя формулы из шаблона (первый лист)."""
    date_str = report_date.strftime("%d.%m.%Y")
    formula_row = _get_formula_template_row(template_ws)  # образец формул из шаблона
    next_row = ws.max_row + 1

    for name in names:
        ws.cell(row=next_row, column=1, value=date_str)
        ws.cell(row=next_row, column=2, value=name)
        if formula_row and template_ws.max_column >= 3:
            for col in range(3, template_ws.max_column + 1):
                src_cell = template_ws.cell(row=formula_row, column=col)
                if _is_formula(src_cell.value):
                    new_formula = _copy_formula_row(src_cell.value, formula_row, next_row)
                    ws.cell(row=next_row, column=col, value=new_formula)
        next_row += 1


def _has_data_for_date(ws, report_date: date) -> bool:
    """Проверить, есть ли уже данные за эту дату (избежать дублей)."""
    date_str = report_date.strftime("%d.%m.%Y")
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=1).value == date_str:
            return True
    return False


async def generate_report(poll_date: date, names: list[str]) -> str | None:
    """
    Добавить отчёт за дату в локальный файл Doc/Бадминтон.xlsx.

    Каждый месяц — отдельный лист. Структура по образцу первого листа.
    Returns: путь к файлу при успехе, иначе None.
    """
    path = _get_report_path()
    if not _ensure_report_file(path):
        return None
    try:
        wb = _load_workbook_safe(path)
        ws = _get_or_create_month_sheet(wb, poll_date)
        template_ws = wb.worksheets[0]  # первый лист — образец структуры и формул
        if _has_data_for_date(ws, poll_date):
            logger.info("Данные за %s уже есть в отчёте, пропуск", poll_date)
        else:
            _append_report_data(ws, template_ws, poll_date, names)
        wb.save(path)
        logger.info("Отчёт обновлён: %s, добавлено %d чел.", poll_date, len(names))
        return path
    except Exception as e:
        logger.exception("Ошибка формирования отчёта: %s", e)
        return None


def get_report_path() -> str | None:
    """Путь к файлу отчёта, если он существует."""
    from pathlib import Path
    p = Path(_get_report_path())
    return str(p) if p.exists() else None
