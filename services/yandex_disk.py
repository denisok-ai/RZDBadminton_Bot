"""
@file: yandex_disk.py
@description: Работа с Яндекс.Диск — скачивание шаблона, заполнение, загрузка
@dependencies: yadisk, openpyxl, aiohttp (aiogram), config
@created: 2025-02-25
"""

import asyncio
import logging
from pathlib import Path

import aiohttp
from openpyxl import load_workbook

from config import get_settings

logger = logging.getLogger("rzdbadminton")

MAX_RETRIES = 3


async def _download_public_file(url: str, save_path: Path) -> bool:
    """Скачать файл по публичной ссылке Яндекс.Диск."""
    api_url = "https://cloud-api.yandex.net/v1/disk/public/resources/download"
    params = {"public_key": url}
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(api_url, params=params, timeout=aiohttp.ClientTimeout(total=30)) as resp:
                resp.raise_for_status()
                data = await resp.json()
            href = data.get("href")
            if not href:
                logger.error("Нет ссылки на скачивание в ответе API")
                return False
            async with session.get(href, timeout=aiohttp.ClientTimeout(total=60)) as dl:
                dl.raise_for_status()
                content = await dl.read()
        save_path.parent.mkdir(parents=True, exist_ok=True)
        save_path.write_bytes(content)
        return True
    except Exception as e:
        logger.exception("Ошибка скачивания шаблона: %s", e)
        return False


def _fill_excel_with_names(file_path: Path, names: list[str]) -> bool:
    """
    Заполнить Excel списком имён.
    Первый лист, колонка A, начиная со строки 2.
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        for i, name in enumerate(names, start=2):
            ws.cell(row=i, column=1, value=name)
        wb.save(file_path)
        return True
    except Exception as e:
        logger.exception("Ошибка заполнения Excel: %s", e)
        return False


def _upload_to_disk(local_path: Path, remote_path: str) -> bool:
    """Загрузить файл на Яндекс.Диск."""
    try:
        import yadisk

        client = yadisk.Client(token=get_settings().yandex_disk_token)
        parent = "/" + "/".join(remote_path.strip("/").split("/")[:-1])
        if parent != "/":
            try:
                client.mkdir(parent)
            except yadisk.exceptions.PathExistsError:
                pass
        client.upload(str(local_path), remote_path, overwrite=True)
        return True
    except Exception as e:
        logger.exception("Ошибка загрузки на Яндекс.Диск: %s", e)
        return False


async def generate_report(
    poll_date,
    names: list[str],
) -> bool:
    """
    Скачать шаблон, заполнить именами, загрузить обратно.

    Args:
        poll_date: дата для отчёта
        names: список имён (ФИ или username)

    Returns:
        True при успехе.
    """
    settings = get_settings()
    template_url = settings.report_template_url
    upload_path = f"{settings.report_upload_path}_{poll_date}.xlsx"
    tmp_dir = Path("data/reports")
    tmp_dir.mkdir(parents=True, exist_ok=True)
    local_path = tmp_dir / f"report_{poll_date}.xlsx"

    for attempt in range(1, MAX_RETRIES + 1):
        if await _download_public_file(template_url, local_path):
            break
        if attempt < MAX_RETRIES:
            await asyncio.sleep(2 ** attempt)
    else:
        logger.error("Не удалось скачать шаблон после %d попыток", MAX_RETRIES)
        return False

    if not _fill_excel_with_names(local_path, names):
        return False

    for attempt in range(1, MAX_RETRIES + 1):
        if _upload_to_disk(local_path, upload_path):
            logger.info("Отчёт загружен: %s", upload_path)
            return True
        if attempt < MAX_RETRIES:
            await asyncio.sleep(2 ** attempt)

    logger.error("Не удалось загрузить отчёт после %d попыток", MAX_RETRIES)
    return False
